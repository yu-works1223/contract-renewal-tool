import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# CSVを読み込む
df = pd.read_csv("tenants.csv")

# datetime型に変換
df["contract_start"] = pd.to_datetime(df["contract_start"])

# 更新日（2年後）を計算
df["renewal_date"] = df["contract_start"] + pd.DateOffset(years=2)

# 案内日（更新日の3ヶ月前）を計算
df["notice_date"] = df["renewal_date"] - pd.DateOffset(months=3)

# 日付を見やすい形式に変換（00:00:00を消す）
df["contract_start"] = df["contract_start"].dt.strftime("%Y-%m-%d")
df["renewal_date"] = df["renewal_date"].dt.strftime("%Y-%m-%d")
df["notice_date"] = df["notice_date"].dt.strftime("%Y-%m-%d")

# 更新年・更新月を追加
df["renewal_year"] = pd.to_datetime(df["renewal_date"]).dt.year
df["renewal_month"] = pd.to_datetime(df["renewal_date"]).dt.month

# 月別集計
summary = df.groupby(["renewal_year", "renewal_month"]).size().reset_index(name="count")

# Excelに書き出す
with pd.ExcelWriter("契約更新リスト.xlsx", engine="openpyxl") as writer:
    df[["tenant_id", "tenant_name", "room_number", "contract_start", "notice_date", "renewal_date"]].to_excel(
        writer, sheet_name="一覧", index=False
    )
    summary.to_excel(writer, sheet_name="集計", index=False)

# 書式設定
wb = load_workbook("契約更新リスト.xlsx")

for sheet_name in ["一覧", "集計"]:
    ws = wb[sheet_name]

    # ヘッダーの書式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 列幅を自動調整
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

wb.save("契約更新リスト.xlsx")
print("=== 完了 ===")
print("契約更新リスト.xlsx を出力しました！")
