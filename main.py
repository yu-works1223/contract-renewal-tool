import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ===========================
# 1. CSVを読み込む
# ===========================
df = pd.read_csv("tenants.csv")

# ===========================
# 2. 日付計算
# ===========================
df["contract_start"] = pd.to_datetime(df["contract_start"])
df["renewal_date"] = df["contract_start"] + pd.DateOffset(years=2)
df["notice_date"] = df["renewal_date"] - pd.DateOffset(months=3)

# 満了日までの残日数
today = pd.Timestamp(date.today())
df["days_until_renewal"] = (df["renewal_date"] - today).dt.days

# ===========================
# 3. 集計
# ===========================
df["renewal_year"] = df["renewal_date"].dt.year
df["renewal_month"] = df["renewal_date"].dt.month

summary = df.groupby(["renewal_year", "renewal_month"]).size().reset_index(name="count")

rent_summary = df.groupby("building_name")["rent"].sum().reset_index()
rent_summary.columns = ["物件名", "賃料合計"]

# ===========================
# 4. 日付を見やすい形式に変換
# ===========================
df["contract_start"] = df["contract_start"].dt.strftime("%Y-%m-%d")
df["renewal_date"] = df["renewal_date"].dt.strftime("%Y-%m-%d")
df["notice_date"] = df["notice_date"].dt.strftime("%Y-%m-%d")

# ===========================
# 5. Excelに出力
# ===========================
output_file = "契約更新リスト.xlsx"

output_columns = [
    "tenant_id", "tenant_name", "building_name", "room_number",
    "floor_plan", "area_sqm", "contract_start", "notice_date",
    "renewal_date", "days_until_renewal", "rent", "common_fee",
    "deposit", "contract_type", "status"
]

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df[output_columns].to_excel(writer, sheet_name="一覧", index=False)
    summary.to_excel(writer, sheet_name="更新月別集計", index=False)
    rent_summary.to_excel(writer, sheet_name="物件別賃料集計", index=False)

# ===========================
# 6. 書式設定
# ===========================
wb = load_workbook(output_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 4

# 一覧シートにオートフィルター＆残日数ハイライト
ws_list = wb["一覧"]
ws_list.auto_filter.ref = ws_list.dimensions

# days_until_renewalの列番号を取得
header = [cell.value for cell in ws_list[1]]
days_col = header.index("days_until_renewal") + 1

red_fill = PatternFill(fill_type="solid", fgColor="FF9999")    # 30日以内
yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF99") # 90日以内

for row in ws_list.iter_rows(min_row=2):
    days_cell = row[days_col - 1]
    try:
        days = int(days_cell.value)
        if days <= 30:
            for cell in row:
                cell.fill = red_fill
        elif days <= 90:
            for cell in row:
                cell.fill = yellow_fill
    except:
        pass

wb.save(output_file)

print("=== 完了 ===")
print(f"{output_file} を出力しました！")
print(f"シート：一覧 / 更新月別集計 / 物件別賃料集計")
print(f"・オートフィルター付き（満了日近い順・物件順で並び替え可）")
print(f"・残30日以内：赤ハイライト / 残90日以内：黄ハイライト")
