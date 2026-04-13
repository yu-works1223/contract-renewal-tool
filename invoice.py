import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===========================
# 1. CSVを読み込む
# ===========================
df = pd.read_csv("tenants.csv")

# ===========================
# 2. 日付計算・対象者を絞り込む
# ===========================
df["contract_start"] = pd.to_datetime(df["contract_start"])
df["renewal_date"] = df["contract_start"] + pd.DateOffset(years=2)
df["notice_date"] = df["renewal_date"] - pd.DateOffset(months=3)
df["payment_due"] = df["renewal_date"] - pd.DateOffset(months=1)

today = pd.Timestamp(date.today())
df["days_until_renewal"] = (df["renewal_date"] - today).dt.days

targets = df[df["days_until_renewal"] <= 90].copy()
print(f"対象件数：{len(targets)}件")

# ===========================
# 振込口座情報（実際の口座に変更してください）
# ===========================
BANK_NAME = "○○銀行"
BRANCH_NAME = "△△支店"
ACCOUNT_TYPE = "普通"
ACCOUNT_NUMBER = "1234567"
ACCOUNT_HOLDER = "カ）○○管理"

COMPANY_NAME = "○○不動産管理株式会社"
COMPANY_TEL = "03-XXXX-XXXX"
COMPANY_ADDRESS = "東京都○○区○○1-2-3"

# ===========================
# 3. 請求書Excelを作成
# ===========================
wb = Workbook()
wb.remove(wb.active)

thin = Side(style="thin")
medium = Side(style="medium")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)
border_bottom_medium = Border(bottom=medium)

for _, row in targets.iterrows():
    sheet_name = f"{row['building_name']}_{row['room_number']}_{row['tenant_name']}"
    sheet_name = sheet_name[:31]
    ws = wb.create_sheet(title=sheet_name)

    # A4設定
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.9
    ws.page_margins.bottom = 0.9

    # 列幅設定
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16

    # 金額計算
    rent = int(row["rent"])
    fee = int(rent * 0.25)
    tax = int(fee * 0.1)
    total = rent + fee + tax

    renewal_date_str = row["renewal_date"].strftime("%Y年%m月%d日")
    payment_due_str = row["payment_due"].strftime("%Y年%m月%d日")
    issue_date_str = today.strftime("%Y年%m月%d日")

    # ===========================
    # タイトル
    # ===========================
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:E2")
    ws["B2"] = "更　新　料　請　求　書"
    ws["B2"].font = Font(size=20, bold=True)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    # 発行日
    ws.row_dimensions[3].height = 18
    ws.merge_cells("B3:E3")
    ws["B3"] = f"発行日：{issue_date_str}"
    ws["B3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B3"].font = Font(size=10)

    # ===========================
    # 宛先・発行者
    # ===========================
    ws.row_dimensions[5].height = 22
    ws["B5"] = f"{row['tenant_name']}　様"
    ws["B5"].font = Font(size=14, bold=True)
    ws["B5"].alignment = Alignment(vertical="center")

    ws.row_dimensions[6].height = 16
    ws["D6"] = COMPANY_NAME
    ws["D6"].font = Font(size=10, bold=True)
    ws["D6"].alignment = Alignment(horizontal="left")

    ws.row_dimensions[7].height = 14
    ws["D7"] = COMPANY_ADDRESS
    ws["D7"].font = Font(size=9, color="444444")

    ws.row_dimensions[8].height = 14
    ws["D8"] = f"TEL：{COMPANY_TEL}"
    ws["D8"].font = Font(size=9, color="444444")

    # 区切り線
    ws.row_dimensions[9].height = 6
    for col in ["B", "C", "D", "E"]:
        ws[f"{col}9"].border = Border(bottom=medium)

    # ===========================
    # 物件情報
    # ===========================
    ws.row_dimensions[10].height = 6
    ws.row_dimensions[11].height = 18

    info = [
        ("物件名", row["building_name"]),
        ("部屋番号", str(row["room_number"])),
        ("契約更新日", renewal_date_str),
        ("お支払期限", payment_due_str),
    ]

    for i, (label, value) in enumerate(info, start=11):
        ws.row_dimensions[i].height = 18
        ws[f"B{i}"] = label
        ws[f"B{i}"].font = Font(bold=True, size=10)
        ws[f"B{i}"].alignment = Alignment(vertical="center")
        ws[f"C{i}"] = value
        ws[f"C{i}"].font = Font(size=10)
        ws[f"C{i}"].alignment = Alignment(vertical="center")

    # ===========================
    # 請求明細
    # ===========================
    ws.row_dimensions[16].height = 6
    ws.row_dimensions[17].height = 20

    # ヘッダー
    headers = ["項目", "金額（円）", "備考"]
    cols = ["B", "C", "D"]
    for col, header in zip(cols, headers):
        ws[f"{col}17"] = header
        ws[f"{col}17"].font = Font(bold=True, color="FFFFFF", size=10)
        ws[f"{col}17"].fill = PatternFill(fill_type="solid", fgColor="4472C4")
        ws[f"{col}17"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col}17"].border = border_thin

    # 明細
    items = [
        ("更新料", rent, "賃料1ヶ月分（非課税）"),
        ("更新事務手数料", fee, "賃料×0.25（税抜）"),
        ("消費税", tax, "手数料に対する10%"),
    ]

    for i, (name, amount, note) in enumerate(items, start=18):
        ws.row_dimensions[i].height = 18
        ws[f"B{i}"] = name
        ws[f"C{i}"] = amount
        ws[f"D{i}"] = note
        ws[f"C{i}"].number_format = "#,##0"
        ws[f"C{i}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"B{i}"].alignment = Alignment(vertical="center")
        ws[f"D{i}"].alignment = Alignment(vertical="center")
        for col in ["B", "C", "D"]:
            ws[f"{col}{i}"].border = border_thin
            ws[f"{col}{i}"].font = Font(size=10)

    # 合計行
    ws.row_dimensions[21].height = 22
    ws["B21"] = "合　計"
    ws["C21"] = total
    ws["C21"].number_format = "#,##0"
    ws["B21"].font = Font(bold=True, size=12)
    ws["C21"].font = Font(bold=True, size=12)
    ws["B21"].fill = PatternFill(fill_type="solid", fgColor="DDEEFF")
    ws["C21"].fill = PatternFill(fill_type="solid", fgColor="DDEEFF")
    ws["B21"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C21"].alignment = Alignment(horizontal="right", vertical="center")
    for col in ["B", "C", "D"]:
        ws[f"{col}21"].border = border_thin

    # ===========================
    # 振込口座
    # ===========================
    ws.row_dimensions[23].height = 6
    ws.row_dimensions[24].height = 20

    ws.merge_cells("B24:E24")
    ws["B24"] = "■ お振込先"
    ws["B24"].font = Font(bold=True, size=11)
    ws["B24"].fill = PatternFill(fill_type="solid", fgColor="EEF4FF")
    ws["B24"].alignment = Alignment(vertical="center")
    ws["B24"].border = Border(bottom=thin)

    bank_info = [
        ("金融機関", f"{BANK_NAME}　{BRANCH_NAME}"),
        ("口座種別", ACCOUNT_TYPE),
        ("口座番号", ACCOUNT_NUMBER),
        ("口座名義", ACCOUNT_HOLDER),
    ]

    for i, (label, value) in enumerate(bank_info, start=25):
        ws.row_dimensions[i].height = 18
        ws[f"B{i}"] = label
        ws[f"C{i}"] = value
        ws[f"B{i}"].font = Font(bold=True, size=10)
        ws[f"C{i}"].font = Font(size=10)
        ws[f"B{i}"].alignment = Alignment(vertical="center")
        ws[f"C{i}"].alignment = Alignment(vertical="center")

    # 注意書き
    ws.row_dimensions[30].height = 16
    ws.merge_cells("B30:E30")
    ws["B30"] = "※ 振込手数料はご負担ください。ご不明点はお気軽にお問い合わせください。"
    ws["B30"].font = Font(size=9, color="FF0000")
    ws["B30"].alignment = Alignment(vertical="center")

wb.save("更新料請求書.xlsx")
print("=== 完了 ===")
print("更新料請求書.xlsx を出力しました！")
